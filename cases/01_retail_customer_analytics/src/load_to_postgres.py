"""Load UCI Online Retail II into PostgreSQL with an auditable trail.

The password is requested through getpass and is never written to disk.
Rows are streamed from Excel to PostgreSQL COPY in bounded batches.
"""

from __future__ import annotations

import argparse
import csv
import getpass
import hashlib
import io
import json
import os
import sys
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable

import psycopg2
from openpyxl import load_workbook


LOADER_VERSION = "1.0.0"
EXPECTED_HEADERS = [
    "Invoice",
    "StockCode",
    "Description",
    "Quantity",
    "InvoiceDate",
    "Price",
    "Customer ID",
    "Country",
]
EXPECTED_SHEETS = ["Year 2009-2010", "Year 2010-2011"]
COPY_COLUMNS = (
    "invoice, stock_code, description, quantity, invoice_date, price, "
    "customer_id, country, source_sheet, source_row_number"
)


def parse_args() -> argparse.Namespace:
    case_dir = Path(__file__).resolve().parents[1]
    parser = argparse.ArgumentParser(
        description="Stream Online Retail II from Excel into PostgreSQL."
    )
    parser.add_argument(
        "--source",
        type=Path,
        default=case_dir / "data" / "raw" / "online_retail_II.xlsx",
    )
    parser.add_argument("--host", default="localhost")
    parser.add_argument("--port", type=int, default=5432)
    parser.add_argument("--database", default="golden_data_lab")
    parser.add_argument("--user", default="postgres")
    parser.add_argument("--batch-size", type=int, default=10_000)
    parser.add_argument(
        "--replace",
        action="store_true",
        help="Explicitly truncate an existing raw load before loading again.",
    )
    parser.add_argument(
        "--profile-only",
        action="store_true",
        help="Inspect source metadata without connecting to PostgreSQL.",
    )
    return parser.parse_args()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def normalize_integer(value: Any, field_name: str, sheet: str, row_number: int) -> int | None:
    if value is None:
        return None
    numeric = Decimal(str(value))
    if numeric != numeric.to_integral_value():
        raise ValueError(
            f"{sheet} row {row_number}: {field_name}={value!r} is not an integer"
        )
    return int(numeric)


def normalize_datetime(value: Any, sheet: str, row_number: int) -> str | None:
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat(sep=" ")
    raise ValueError(
        f"{sheet} row {row_number}: InvoiceDate={value!r} is not a date/time"
    )


def normalize_decimal(value: Any) -> str | None:
    if value is None:
        return None
    return format(Decimal(str(value)), "f")


def inspect_workbook(path: Path) -> tuple[list[str], dict[str, int]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if workbook.sheetnames != EXPECTED_SHEETS:
            raise ValueError(
                f"Unexpected sheets: {workbook.sheetnames}; expected {EXPECTED_SHEETS}"
            )

        row_counts: dict[str, int] = {}
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
            if headers != EXPECTED_HEADERS:
                raise ValueError(
                    f"Unexpected headers in {sheet_name}: {headers}; "
                    f"expected {EXPECTED_HEADERS}"
                )
            row_counts[sheet_name] = worksheet.max_row - 1
        return workbook.sheetnames, row_counts
    finally:
        workbook.close()


def iter_normalized_rows(
    worksheet: Any, sheet_name: str
) -> Iterable[tuple[Any, ...]]:
    for row_number, values in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if len(values) != len(EXPECTED_HEADERS):
            raise ValueError(
                f"{sheet_name} row {row_number}: expected 8 fields, got {len(values)}"
            )
        yield (
            normalize_text(values[0]),
            normalize_text(values[1]),
            normalize_text(values[2]),
            normalize_integer(values[3], "Quantity", sheet_name, row_number),
            normalize_datetime(values[4], sheet_name, row_number),
            normalize_decimal(values[5]),
            normalize_integer(values[6], "Customer ID", sheet_name, row_number),
            normalize_text(values[7]),
            sheet_name,
            row_number,
        )


def copy_batch(cursor: Any, rows: list[tuple[Any, ...]]) -> None:
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerows(rows)
    buffer.seek(0)
    cursor.copy_expert(
        f"""
        COPY retail.online_retail_raw ({COPY_COLUMNS})
        FROM STDIN WITH (FORMAT CSV, NULL '')
        """,
        buffer,
    )


def load_sheet(cursor: Any, source: Path, sheet_name: str, batch_size: int) -> int:
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        batch: list[tuple[Any, ...]] = []
        loaded = 0
        for row in iter_normalized_rows(worksheet, sheet_name):
            batch.append(row)
            if len(batch) >= batch_size:
                copy_batch(cursor, batch)
                loaded += len(batch)
                print(f"  {sheet_name}: {loaded:,} rows copied", flush=True)
                batch.clear()
        if batch:
            copy_batch(cursor, batch)
            loaded += len(batch)
        print(f"  {sheet_name}: {loaded:,} rows copied (complete)", flush=True)
        return loaded
    finally:
        workbook.close()


def ensure_empty_target(cursor: Any, replace: bool) -> None:
    cursor.execute("SELECT count(*) FROM retail.online_retail_raw")
    existing_rows = cursor.fetchone()[0]
    if existing_rows == 0:
        return
    if not replace:
        raise RuntimeError(
            f"Target already contains {existing_rows:,} rows. "
            "Stop to avoid duplicate loading; use --replace only after reviewing the target."
        )
    cursor.execute("TRUNCATE TABLE retail.online_retail_raw, retail.load_audit RESTART IDENTITY")


def main() -> int:
    args = parse_args()
    source = args.source.resolve()
    if not source.is_file():
        raise FileNotFoundError(source)
    if args.batch_size <= 0:
        raise ValueError("--batch-size must be positive")

    sheets, sheet_row_counts = inspect_workbook(source)
    source_sha256 = sha256_file(source)
    source_size = source.stat().st_size
    expected_rows = sum(sheet_row_counts.values())

    print("Source verification")
    print(f"  file: {source}")
    print(f"  SHA-256: {source_sha256}")
    print(f"  size: {source_size:,} bytes")
    for sheet_name, row_count in sheet_row_counts.items():
        print(f"  {sheet_name}: {row_count:,} data rows")
    print(f"  expected total: {expected_rows:,} rows")

    if args.profile_only:
        return 0

    password = os.environ.get("PGPASSWORD") or getpass.getpass(
        f"PostgreSQL password for {args.user}: "
    )
    started_at = datetime.now(timezone.utc)
    connection = psycopg2.connect(
        host=args.host,
        port=args.port,
        dbname=args.database,
        user=args.user,
        password=password,
    )
    try:
        with connection:
            with connection.cursor() as cursor:
                ensure_empty_target(cursor, args.replace)
                loaded_by_sheet = {
                    sheet: load_sheet(cursor, source, sheet, args.batch_size)
                    for sheet in sheets
                }
                loaded_rows = sum(loaded_by_sheet.values())
                if loaded_by_sheet != sheet_row_counts:
                    raise RuntimeError(
                        f"Sheet reconciliation failed: loaded={loaded_by_sheet}, "
                        f"expected={sheet_row_counts}"
                    )
                if loaded_rows != expected_rows:
                    raise RuntimeError(
                        f"Total reconciliation failed: loaded={loaded_rows:,}, "
                        f"expected={expected_rows:,}"
                    )

                finished_at = datetime.now(timezone.utc)
                cursor.execute(
                    """
                    INSERT INTO retail.load_audit (
                        source_file_name, source_sha256, source_size_bytes,
                        expected_row_count, loaded_row_count, sheet_row_counts,
                        started_at, finished_at, database_user, loader_version, status
                    )
                    VALUES (%s, %s, %s, %s, %s, %s::jsonb, %s, %s,
                            current_user, %s, 'success')
                    RETURNING load_id, database_user
                    """,
                    (
                        source.name,
                        source_sha256,
                        source_size,
                        expected_rows,
                        loaded_rows,
                        json.dumps(loaded_by_sheet, ensure_ascii=False),
                        started_at,
                        finished_at,
                        LOADER_VERSION,
                    ),
                )
                load_id, database_user = cursor.fetchone()
        print("Load committed successfully")
        print(f"  load_id: {load_id}")
        print(f"  database user: {database_user}")
        print(f"  loaded total: {loaded_rows:,} rows")
        print("Next: run sql/02_validate_raw_load.sql and save its output.")
        return 0
    finally:
        connection.close()


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"LOAD FAILED: {exc}", file=sys.stderr)
        raise SystemExit(1)
